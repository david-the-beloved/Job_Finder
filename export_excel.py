# ──────────────────────────────────────────────────────────────
# export_excel.py -- Export jobs from SQLite to Excel
# ──────────────────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

from database import get_jobs_for_export, get_all_jobs, get_stats
from config import EXCEL_PATH, MIN_QUALITY_SCORE

# ── Colour palette ────────────────────────────────────────────
C = {
    "header_bg":      "1A3C6E",
    "header_fg":      "FFFFFF",
    "row_even":       "F2F7FF",
    "row_odd":        "FFFFFF",
    "remote_yes":     "D4EDDA",
    "remote_hybrid":  "FFF3CD",
    "remote_no":      "F8D7DA",
    "scam":           "FF4444",
    "verified_yes":   "28A745",
    "verified_no":    "DC3545",
    "unverified":     "888888",
    "score_high":     "1E7E34",   # 8-10
    "score_mid":      "856404",   # 5-7
    "score_low":      "721C24",   # 1-4
    "source_indeed":  "2557A7",
    "source_remote":  "0D8050",
    "source_hn":      "FF6600",
    "source_google":  "4285F4",
    "source_other":   "666666",
    "summary_bg":     "E8EEF7",
    "scams_tab_bg":   "FFEBEE",
}

COLUMNS = [
    # (Header label,  col width)
    ("ID",            8),
    ("Source",       16),
    ("Job Title",    36),
    ("Company",      22),
    ("Location",     18),
    ("Salary",       18),
    ("Remote",       10),
    ("Tech Stack",   28),
    ("Score",         8),
    ("Date Posted",  13),
    ("Date Found",   13),
    ("Link",         14),
    ("Verified",     11),
    ("Description",  55),
]

thin   = Side(style="thin",   color="DDDDDD")
thick  = Side(style="medium", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
H_BDR  = Border(left=thick, right=thick, top=thick, bottom=thick)


def _source_color(source: str) -> str:
    s = source.lower()
    if "indeed"   in s: return C["source_indeed"]
    if "remoteok" in s: return C["source_remote"]
    if "hacker"   in s: return C["source_hn"]
    if "google"   in s: return C["source_google"]
    return C["source_other"]


def _score_color(score) -> str:
    if score is None: return C["unverified"]
    if score >= 8:    return C["score_high"]
    if score >= 5:    return C["score_mid"]
    return C["score_low"]


def _write_header(ws, columns=None):
    cols = columns or COLUMNS
    fill = PatternFill("solid", fgColor=C["header_bg"])
    for col, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(name="Arial", size=11, bold=True, color=C["header_fg"])
        cell.fill      = fill
        cell.border    = H_BDR
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _write_job_row(ws, row_num: int, job: dict):
    is_even   = (row_num % 2 == 0)
    base_fill = PatternFill("solid", fgColor=C["row_even"] if is_even else C["row_odd"])

    score   = job.get("quality_score")
    verified = job.get("verified")  # 1, 0, or None

    verified_label = "✓" if verified == 1 else ("✗" if verified == 0 else "—")

    values = [
        job.get("id",          ""),
        job.get("source",      ""),
        job.get("title",       ""),
        job.get("company",     ""),
        job.get("location",    ""),
        job.get("salary",      ""),
        job.get("remote",      ""),
        job.get("tech_stack",  "") or "",
        score if score is not None else "—",
        job.get("date_posted", ""),
        (job.get("date_found", "") or "")[:10],
        job.get("url",         ""),
        verified_label,
        (job.get("description", "") or "")[:300],
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border    = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=(col == 14))
        cell.font      = Font(name="Arial", size=10)
        cell.fill      = base_fill

        # Source -- coloured text
        if col == 2:
            cell.font = Font(name="Arial", size=10, bold=True,
                             color=_source_color(str(val)))

        # Remote pill
        if col == 7:
            v  = str(val)
            bg = (C["remote_yes"]    if v == "Yes"
                  else C["remote_hybrid"] if v == "Hybrid"
                  else C["remote_no"])
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Arial", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="top")

        # Quality score -- colour by value
        if col == 9:
            cell.font      = Font(name="Arial", size=10, bold=True,
                                  color=_score_color(score))
            cell.alignment = Alignment(horizontal="center", vertical="top")

        # Hyperlink
        if col == 12 and str(val).startswith("http"):
            cell.hyperlink = str(val)
            cell.value     = "View Job"
            cell.font      = Font(name="Arial", size=10,
                                  color="0A66C2", underline="single")

        # Verified badge
        if col == 13:
            color = (C["verified_yes"] if val == "✓"
                     else C["verified_no"] if val == "✗"
                     else C["unverified"])
            cell.font      = Font(name="Arial", size=11, bold=True, color=color)
            cell.alignment = Alignment(horizontal="center", vertical="top")


def _write_scams_sheet(wb, all_jobs: list):
    """Separate tab for scam/low-quality jobs so they're not lost."""
    scams = [j for j in all_jobs if j.get("scam_flag")]
    if not scams:
        return

    if "Scam Flags" in wb.sheetnames:
        del wb["Scam Flags"]
    ws = wb.create_sheet("Scam Flags")

    scam_cols = [
        ("Title",     36), ("Company", 22), ("Source", 16),
        ("Scam Reason", 45), ("Score", 8), ("URL", 30),
    ]
    _write_header(ws, scam_cols)

    for i, job in enumerate(scams):
        r = i + 2
        row_fill = PatternFill("solid", fgColor="FFF0F0" if i % 2 == 0 else "FFFFFF")
        for col, val in enumerate([
            job.get("title", ""),
            job.get("company", ""),
            job.get("source", ""),
            job.get("scam_flag", ""),
            job.get("quality_score", ""),
            job.get("url", ""),
        ], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            cell.fill   = row_fill
            cell.font   = Font(name="Arial", size=10,
                               color=(C["scam"] if col == 4 else "000000"))
            cell.alignment = Alignment(vertical="top")


def _write_summary_sheet(wb, stats: dict, ai_stats: dict):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)

    def hdr(r, text, size=11, bold=False, color="000000"):
        c = ws.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", size=size, bold=bold, color=color)
        c.alignment = Alignment(vertical="center")

    def kv(r, label, value, bold_val=False, val_color="000000"):
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        c1.font = Font(name="Arial", size=11)
        c2.font = Font(name="Arial", size=11, bold=bold_val, color=val_color)
        c1.alignment = c2.alignment = Alignment(vertical="center")

    grey = PatternFill("solid", fgColor=C["summary_bg"])

    hdr(1, "Job Scraper -- Summary", size=15, bold=True, color=C["header_bg"])
    hdr(2, f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", color="888888")

    ws.append([])
    hdr(4, "Collection Stats", size=12, bold=True)
    kv(5, "Total jobs in database", stats.get("total", 0),        bold_val=True)
    kv(6, "Added today",            stats.get("today", 0),        bold_val=True, val_color=C["score_high"])
    kv(7, "Remote jobs",            stats.get("remote", 0),       bold_val=True)
    kv(8, "Scams flagged",          ai_stats.get("scams_total", 0), bold_val=True, val_color=C["scam"])
    kv(9, "Unverified (pending AI)",ai_stats.get("unverified", 0), bold_val=True, val_color=C["unverified"])

    ws.append([])
    hdr(11, "AI Verification", size=12, bold=True)
    kv(12, "Verified jobs",         ai_stats.get("verified_total", 0), bold_val=True)
    kv(13, "Avg quality score",     ai_stats.get("avg_score", "--"),   bold_val=True)
    kv(14, "API calls today",       ai_stats.get("calls_today", 0))
    kv(15, "Est. cost today",       f"${ai_stats.get('cost_today', 0):.4f}")

    ws.append([])
    hdr(17, "Jobs by Source", size=12, bold=True)
    for i, (source, count) in enumerate(stats.get("sources", {}).items(), 18):
        kv(i, source, count)
        if i % 2 == 0:
            for c in [1, 2]:
                ws.cell(row=i, column=c).fill = grey

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def _get_ai_stats() -> dict:
    """Pull AI-related stats from the DB for the summary sheet."""
    from database import get_connection
    conn = get_connection()
    scams_total     = conn.execute("SELECT COUNT(*) FROM jobs WHERE scam_flag != '' AND scam_flag IS NOT NULL").fetchone()[0]
    verified_total  = conn.execute("SELECT COUNT(*) FROM jobs WHERE verified = 1").fetchone()[0]
    unverified      = conn.execute("SELECT COUNT(*) FROM jobs WHERE verified IS NULL").fetchone()[0]
    avg_row         = conn.execute("SELECT AVG(quality_score) FROM jobs WHERE quality_score IS NOT NULL").fetchone()[0]
    avg_score       = round(avg_row, 1) if avg_row else "--"
    conn.close()

    # Read today's cost from quota file
    calls_today = 0
    cost_today  = 0.0
    try:
        import json
        from config import GEMINI_QUOTA_FILE
        if os.path.exists(GEMINI_QUOTA_FILE):
            with open(GEMINI_QUOTA_FILE) as f:
                q = json.load(f)
            from datetime import date
            if q.get("date") == str(date.today()):
                calls_today = q.get("calls", 0)
                cost_today  = q.get("cost_usd", 0.0)
    except Exception:
        pass

    return {
        "scams_total":    scams_total,
        "verified_total": verified_total,
        "unverified":     unverified,
        "avg_score":      avg_score,
        "calls_today":    calls_today,
        "cost_today":     cost_today,
    }


def export(jobs: list = None) -> str:
    """
    Export clean jobs to Excel. Scams get their own tab.
    Returns the file path.
    """
    jobs     = jobs or get_jobs_for_export(min_quality=MIN_QUALITY_SCORE)
    all_jobs = get_all_jobs()   # for scam tab
    stats    = get_stats()
    ai_stats = _get_ai_stats()

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Postings"

    _write_header(ws)
    for i, job in enumerate(jobs):
        _write_job_row(ws, i + 2, job)

    _write_scams_sheet(wb, all_jobs)
    _write_summary_sheet(wb, stats, ai_stats)

    wb.active = wb["Summary"]

    os.makedirs(os.path.dirname(EXCEL_PATH) if os.path.dirname(EXCEL_PATH) else ".", exist_ok=True)
    wb.save(EXCEL_PATH)

    scam_count = len([j for j in all_jobs if j.get("scam_flag")])
    print(f"  Excel saved -> {EXCEL_PATH}")
    print(f"  Sheets: Job Postings ({len(jobs)} jobs)  |  Scam Flags ({scam_count})  |  Summary")
    return EXCEL_PATH


if __name__ == "__main__":
    export()
