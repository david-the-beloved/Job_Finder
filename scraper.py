"""
Job Scraper — Pulls from free RSS/API sources and saves to Excel.
Sources: Google Alerts RSS, Indeed RSS, RemoteOK API, HN Who's Hiring

HOW TO RUN:
  python scraper.py

REQUIREMENTS:
  pip install requests feedparser openpyxl python-dateutil

SETUP:
  1. Replace GOOGLE_ALERT_URLS with your own Google Alert RSS links
  2. Edit INDEED_SEARCHES to match the roles you want
  3. (Optional) Add your Claude API key to AI_VERIFY for job verification
  4. Run the script — jobs_output.xlsx will be created/updated
"""

import requests
import xml.etree.ElementTree as ET
import json
import hashlib
import os
import re
from datetime import datetime
from urllib.parse import urlencode

# ─────────────────────────────────────────────
# CONFIGURATION — Edit these to match your needs
# ─────────────────────────────────────────────

# Paste your Google Alert RSS URLs here (go to google.com/alerts, set delivery to RSS)
GOOGLE_ALERT_URLS = [
    # Example — replace with your real Google Alert RSS URLs:
    # "https://www.google.com/alerts/feeds/YOUR_ID/YOUR_ALERT_ID",
]

# Job roles + locations to search on Indeed
INDEED_SEARCHES = [
    {"q": "software engineer", "l": "remote"},
    {"q": "product manager", "l": "remote"},
    {"q": "data scientist", "l": "remote"},
    {"q": "frontend developer", "l": "remote"},
    {"q": "backend developer", "l": "remote"},
]

# Claude API key for AI verification (optional — leave empty to skip)
CLAUDE_API_KEY = ""  # e.g. "sk-ant-..."

# Output file
OUTPUT_FILE = "jobs_output.xlsx"

# ─────────────────────────────────────────────
# SCRAPERS
# ─────────────────────────────────────────────

def parse_rss(url, source_name):
    """Generic RSS parser — works for Google Alerts and Indeed."""
    jobs = []
    try:
        resp = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        root = ET.fromstring(resp.content)

        # Handle both RSS 2.0 and Atom formats
        ns = {"atom": "http://www.w3.org/2005/Atom"}
        items = root.findall(".//item") or root.findall(".//atom:entry", ns)

        for item in items:
            def get(tag, atom_tag=None):
                el = item.find(tag)
                if el is None and atom_tag:
                    el = item.find(atom_tag, ns)
                return el.text.strip() if el is not None and el.text else ""

            title = get("title", "atom:title")
            link  = get("link",  "atom:link")
            desc  = get("description", "atom:summary")
            date  = get("pubDate", "atom:updated")

            if not link:
                link_el = item.find("atom:link", ns)
                if link_el is not None:
                    link = link_el.get("href", "")

            # Clean HTML from description
            desc_clean = re.sub(r"<[^>]+>", " ", desc).strip()
            desc_clean = re.sub(r"\s+", " ", desc_clean)[:500]

            jobs.append({
                "source":      source_name,
                "title":       title,
                "company":     extract_company(title, desc_clean),
                "location":    extract_location(desc_clean),
                "salary":      extract_salary(desc_clean),
                "remote":      is_remote(title + " " + desc_clean),
                "description": desc_clean,
                "url":         link,
                "date_found":  parse_date(date),
                "verified":    "",
                "scam_flag":   "",
                "id":          hashlib.md5(link.encode()).hexdigest()[:12],
            })
    except Exception as e:
        print(f"  ✗ {source_name}: {e}")
    return jobs


def scrape_indeed(query, location):
    """Scrape Indeed via their public RSS feed."""
    params = {"q": query, "l": location, "sort": "date", "limit": 50}
    url = f"https://www.indeed.com/rss?{urlencode(params)}"
    return parse_rss(url, f"Indeed ({query})")


def scrape_remoteok():
    """RemoteOK public JSON API — no key needed."""
    jobs = []
    try:
        resp = requests.get(
            "https://remoteok.com/api",
            timeout=10,
            headers={"User-Agent": "Mozilla/5.0"}
        )
        data = resp.json()
        for item in data[1:]:  # first item is a notice
            if not isinstance(item, dict):
                continue
            desc = re.sub(r"<[^>]+>", " ", item.get("description", ""))[:500]
            jobs.append({
                "source":      "RemoteOK",
                "title":       item.get("position", ""),
                "company":     item.get("company", ""),
                "location":    "Remote",
                "salary":      f"${item.get('salary_min','')}-${item.get('salary_max','')}"
                               if item.get("salary_min") else "",
                "remote":      "Yes",
                "description": desc.strip(),
                "url":         item.get("url", ""),
                "date_found":  parse_date(item.get("date", "")),
                "verified":    "",
                "scam_flag":   "",
                "id":          hashlib.md5(item.get("url","").encode()).hexdigest()[:12],
            })
    except Exception as e:
        print(f"  ✗ RemoteOK: {e}")
    return jobs


def scrape_google_alerts(urls):
    """Scrape all Google Alert RSS feeds."""
    jobs = []
    for i, url in enumerate(urls):
        print(f"  → Google Alert {i+1}/{len(urls)}")
        jobs += parse_rss(url, "Google Alerts")
    return jobs

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def extract_company(title, desc):
    patterns = [r"at ([A-Z][a-zA-Z0-9\s&]+)", r"@ ([A-Z][a-zA-Z0-9\s&]+)"]
    for p in patterns:
        m = re.search(p, title + " " + desc)
        if m:
            return m.group(1).strip()[:50]
    return ""

def extract_location(text):
    m = re.search(
        r"\b(remote|new york|san francisco|london|austin|seattle|chicago|boston|"
        r"los angeles|toronto|berlin|amsterdam|singapore)[^\,\.]{0,20}",
        text, re.IGNORECASE
    )
    return m.group(0).strip().title() if m else ""

def extract_salary(text):
    m = re.search(r"(\$[\d,]+\s*[-–to]+\s*\$[\d,]+|\$[\d,]+[k+/yr]*)", text, re.IGNORECASE)
    return m.group(0) if m else ""

def is_remote(text):
    return "Yes" if re.search(r"\bremote\b", text, re.IGNORECASE) else "No"

def parse_date(raw):
    if not raw:
        return datetime.now().strftime("%Y-%m-%d")
    for fmt in ["%a, %d %b %Y %H:%M:%S %z", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d"]:
        try:
            return datetime.strptime(raw[:25], fmt).strftime("%Y-%m-%d")
        except:
            pass
    return datetime.now().strftime("%Y-%m-%d")

# ─────────────────────────────────────────────
# AI VERIFICATION (Optional — needs Claude API key)
# ─────────────────────────────────────────────

def verify_with_claude(job):
    """Call Claude API to verify job and extract structured fields."""
    if not CLAUDE_API_KEY:
        return job
    try:
        prompt = f"""Analyze this job posting and return ONLY a JSON object with these fields:
- verified: true/false (is this a real, legitimate job posting?)
- scam_flag: "" or short reason if suspicious
- company: company name (improve if missing)
- location: city/remote (improve if missing)
- salary: salary range if mentioned
- remote: "Yes" or "No"

Job title: {job['title']}
Description: {job['description'][:400]}
URL: {job['url']}

Return ONLY valid JSON, no other text."""

        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": CLAUDE_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 300,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=15,
        )
        text = resp.json()["content"][0]["text"]
        data = json.loads(text.strip())
        job.update({k: v for k, v in data.items() if k in job})
    except Exception as e:
        print(f"    AI verify failed: {e}")
    return job

# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

def save_to_excel(jobs, filepath):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    # Load existing or create new
    existing_ids = set()
    if os.path.exists(filepath):
        try:
            wb_existing = load_workbook(filepath)
            ws_existing = wb_existing.active
            for row in ws_existing.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    existing_ids.add(str(row[0]))
            wb_existing.close()
        except:
            pass

    # Filter to only new jobs
    new_jobs = [j for j in jobs if j["id"] not in existing_ids]
    print(f"\n  {len(new_jobs)} new jobs to add (skipping {len(jobs)-len(new_jobs)} duplicates)")

    if not new_jobs and os.path.exists(filepath):
        print("  ✓ No new jobs — Excel file unchanged.")
        return 0

    # Build workbook
    if os.path.exists(filepath) and existing_ids:
        wb = load_workbook(filepath)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Postings"
        start_row = 2
        _write_header(ws)

    # Color palette
    COLORS = {
        "linkedin":     "0A66C2",
        "indeed":       "2557A7",
        "remoteok":     "0D8050",
        "google":       "4285F4",
        "default":      "555555",
        "remote_yes":   "D4EDDA",
        "remote_no":    "F8D7DA",
        "scam":         "FF4444",
        "verified_yes": "28A745",
        "row_even":     "F2F7FF",
        "row_odd":      "FFFFFF",
    }

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, job in enumerate(new_jobs):
        r = start_row + i
        row_color = COLORS["row_even"] if i % 2 == 0 else COLORS["row_odd"]
        fill = PatternFill("solid", fgColor=row_color)

        values = [
            job["id"],
            job["source"],
            job["title"],
            job["company"],
            job["location"],
            job["salary"],
            job["remote"],
            job["date_found"],
            job["url"],
            str(job.get("verified", "")),
            str(job.get("scam_flag", "")),
            job["description"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 12))

            # Source badge color
            if col == 2:
                src = val.lower()
                color = (COLORS["linkedin"] if "linkedin" in src
                         else COLORS["indeed"] if "indeed" in src
                         else COLORS["remoteok"] if "remoteok" in src
                         else COLORS["google"] if "google" in src
                         else COLORS["default"])
                cell.font = Font(name="Arial", size=10, color=color, bold=True)

            # Remote pill
            if col == 7:
                cell.fill = PatternFill("solid", fgColor=(
                    COLORS["remote_yes"] if val == "Yes" else COLORS["remote_no"]
                ))
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")

            # URL as hyperlink
            if col == 9 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.value = "🔗 View Job"
                cell.font = Font(name="Arial", size=10, color="0A66C2",
                                 underline="single")

            # Scam flag
            if col == 11 and val and val.lower() not in ("", "false", "none"):
                cell.font = Font(name="Arial", size=10, color=COLORS["scam"],
                                 bold=True)

            # Verified
            if col == 10 and val:
                color = COLORS["verified_yes"] if val.lower() == "true" else "FF6B6B"
                cell.font = Font(name="Arial", size=10, color=color, bold=True)

    # Summary sheet
    _write_summary_sheet(wb, jobs, new_jobs)

    wb.save(filepath)
    return len(new_jobs)


def _write_header(ws):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    headers = [
        ("ID",          6),
        ("Source",     14),
        ("Job Title",  35),
        ("Company",    22),
        ("Location",   18),
        ("Salary",     16),
        ("Remote",     10),
        ("Date Found", 13),
        ("Link",       14),
        ("Verified",   11),
        ("Scam Flag",  14),
        ("Description",55),
    ]

    hdr_fill = PatternFill("solid", fgColor="1A3C6E")
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L1"


def _write_summary_sheet(wb, all_jobs, new_jobs):
    from openpyxl.styles import Font, PatternFill, Alignment

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)

    # Count by source
    source_counts = {}
    for j in all_jobs:
        s = j["source"].split("(")[0].strip()
        source_counts[s] = source_counts.get(s, 0) + 1

    remote_count = sum(1 for j in all_jobs if j.get("remote") == "Yes")

    rows = [
        ["📊 Job Scraper — Summary", ""],
        ["Last updated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Total jobs collected", len(all_jobs)],
        ["New jobs this run", len(new_jobs)],
        ["Remote jobs", remote_count],
        ["", ""],
        ["Source", "Count"],
    ] + [[s, c] for s, c in sorted(source_counts.items(), key=lambda x: -x[1])]

    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if r == 1 and c == 1:
                cell.font = Font(name="Arial", size=14, bold=True, color="1A3C6E")
            elif r == 7:
                cell.font = Font(name="Arial", size=11, bold=True)
                cell.fill = PatternFill("solid", fgColor="E8EEF7")
            else:
                cell.font = Font(name="Arial", size=11)
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    print("🔍 Job Scraper Starting...\n")
    all_jobs = []

    # 1. Google Alerts RSS
    if GOOGLE_ALERT_URLS:
        print(f"📡 Google Alerts ({len(GOOGLE_ALERT_URLS)} feeds)...")
        all_jobs += scrape_google_alerts(GOOGLE_ALERT_URLS)
    else:
        print("⚠️  No Google Alert URLs set — skipping (add them to GOOGLE_ALERT_URLS)")

    # 2. Indeed RSS
    print(f"\n📡 Indeed ({len(INDEED_SEARCHES)} searches)...")
    for search in INDEED_SEARCHES:
        print(f"  → {search['q']} in {search['l']}")
        all_jobs += scrape_indeed(search["q"], search["l"])

    # 3. RemoteOK
    print("\n📡 RemoteOK API...")
    all_jobs += scrape_remoteok()

    print(f"\n✅ Raw jobs collected: {len(all_jobs)}")

    # 4. AI Verification (if API key set)
    if CLAUDE_API_KEY:
        print(f"\n🤖 AI verifying {len(all_jobs)} jobs...")
        all_jobs = [verify_with_claude(j) for j in all_jobs]

    # 5. Save to Excel
    print(f"\n💾 Saving to {OUTPUT_FILE}...")
    added = save_to_excel(all_jobs, OUTPUT_FILE)
    print(f"\n✅ Done! {added} new jobs saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
